import openpyxl
import csv
# from openpyxl.utils import get_column_letter
from openpyxl.worksheet import worksheet

arr = []
text = "TV Juhtimispult "
text2 = "ТВ пульт "
with open('D:\Kaup24\Samsung121022.csv', newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
    for row in spamreader:
        # row[0] = text + row[0]
        arr.append(row)

category_idA = 4202
# supplier_codeB = [132132133, 132132134, 132132135, 132132136]
# barcodeC = [978064437963, 978064437964, 978064437965, 978064437966]
photo_urlD = ""
# manufacturer_codeE = [4465465, 4465466, 4465467, 4465469]
titleF = "title"
title_lvG = "title_lv"
title_eeH = "title_ee"
title_ruI = "title_ru"
title_fiJ = "title_fi"
long_descriptionK = "long_description"
long_description_lvL = "long_description_lv"
long_description_eeM = "long_description_ee"
long_description_ruN = "long_description_ru"
long_description_fiO = "long_description_fi"
modification_titleP = "modification_title"
modification_title_lvQ = "modification_title_lv"
modification_title_eeR = "modification_title_ee"
modification_title_ruS = "modification_title_ru"
modification_title_fiT = "modification_title_fi"
pack_weightU = 0.2
lengthV = 0.17
widthW = 0.05
heightX = 0.02

filename = "products_for_upload_to_site.xlsx"

book = openpyxl.load_workbook(filename=filename)

# sheet_obj = wb_obj.active
# sheet_obj = wb_obj.worksheets[0]
sheet: worksheet = book["Worksheet"]

# вставить пустую первую строку
# sheet.insert_rows(0)
counts = 0

for i in range(len(arr)):
    arr[i].append(text + str(arr[i][0]))
    arr[i].append(text2 + str(arr[i][0]))

for i in range(3, len(arr) + 2):
    print(i)
    sheet[f"A{i}"].value = category_idA
    sheet[f"B{i}"].value = arr[counts][1]
    # sheet[f"C{i}"].value = barcodeC[counts]
    sheet[f"D{i}"].value = photo_urlD
    # sheet[f"E{i}"].value = manufacturer_codeE[counts]
    sheet[f"H{i}"].value = arr[counts][4]  # titleee
    sheet[f"M{i}"].value = arr[counts][4]  # long description EE
    sheet[f"U{i}"].value = pack_weightU
    sheet[f"V{i}"].value = lengthV
    sheet[f"W{i}"].value = widthW
    sheet[f"X{i}"].value = heightX
    sheet[f"I{i}"].value = arr[counts][5]  # titleru
    sheet[f"N{i}"].value = arr[counts][5]  # long description RU
    # sheet["G4"].value = title_lvG
    # sheet[f"F{i}"].value = arr[o][0]
    # sheet["J4"].value = title_fiJ
    # sheet["K4"].value = long_descriptionK
    # sheet["L4"].value = long_description_lvL
    # sheet["O4"].value = long_description_fiO
    # sheet["P4"].value = modification_titleP
    # sheet["Q4"].value = modification_title_lvQ
    # sheet["R4"].value = modification_title_eeR
    # sheet["S4"].value = modification_title_ruS
    # sheet["T4"].value = modification_title_fiT
    counts += 1
book.save(filename)

# cell_obj = sheet_obj.cell(row=1, column=1)
# sheet_obj['A5'] = 42
# print(cell_obj.value)
# sheet_obj.save("products_with_errors (2).xlsx")


# for row in sheet_obj.values:
#     for cell in row:
#         print(cell)


# for row in sheet_obj:
#     for cell in row:
#         print(cell.font)


# for row in sheet_obj.iter_rows():
#     for cell in row:
#         print(cell.value)
