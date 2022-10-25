import csv
import openpyxl
from openpyxl.worksheet import worksheet

filename = "products_for_upload_to_site.xlsx"
text = "TV Juhtimispult "
text2 = "ТВ пульт "
supplier_file = "exists_supplier_barcodes/supplierProducts.xlsx"
exists_supplier = "exists_supplier_barcodes/existBarcodes.xlsx"


def set_maker():
    """
    создает множество из штрихкодов всех товаров
    :return: dicts_set множество из штрихкодов всех товаров и dicts словарь со всеми товарами
    """
    dicts = {}
    arr = []
    with open('Kaup24.ee_kaup.csv', newline='') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in spamreader:
            arr.append(row)
        for i in arr:
            dicts[i[4]] = [i[0], i[1], i[2], i[3]]
    dicts_set = set(dicts.keys())
    return dicts_set, dicts


def del_set_maker(path):
    """
    создаеи множество из удаляемых щтрихкодов
    :return: это множество из удаляемых щтрихкодов
    """
    barcodes_del = set()
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        # print(str(cell_obj.value))
        barcodes_del.add(str(cell_obj.value))
    return barcodes_del


def sell_dict_full():
    """
    используя два множества получаю готовый список для экспорта товара
    :return: список для экспорта товара
    """
    lst = []
    full_set, dicts = set_maker()
    del_set = del_set_maker(supplier_file) | del_set_maker(exists_supplier)
    finish_set = full_set - del_set
    new_dicts = {}
    for key, value in dicts.items():
        if key in finish_set:
            new_dicts[key] = value
    return new_dicts
    # for key, value in new_dicts.items():
    #     print(key, value)


new_dict = sell_dict_full()

book = openpyxl.load_workbook(filename=filename)
sheet: worksheet = book["Worksheet"]


def name_adding():
    """
    добавление дополнительного названия на русском и эстонском языках
    :return:
    """
    for key, value in new_dict.items():
        value.append(text + value[0])
        value.append(text2 + value[0])


name_adding()
print(new_dict)
